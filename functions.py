from Whatsapp import Whatsapp
from openpyxl import load_workbook
import re
from datetime import datetime as dt
import classes

column_names = [
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
]


def retrieveExcelInfo():
    book = load_workbook("QJS Daily Sales Report  30-Mar-2023.xlsx")
    sheet_names = sorted(book.sheetnames)

    AllSheets = []
    for name in sheet_names:
        AllstoreName = ""
        ListStoreName = ""
        storeFirstName = ""
        instances = 1
        BAVal1 = ""
        BAVal2 = ""

        sheet = book[name]
        if isinstance(sheet["C3"].value, str):
            if sheet["C3"].value.startswith("Store:"):
                dash = sheet["C3"].value.rfind("-")
                if dash > 0:
                    AllstoreName = (sheet["C3"].value[7 : dash - 1]).strip()
                    if AllstoreName.lower().startswith("al"):
                        dash = AllstoreName.rfind("-")
                        if dash > 0:
                            AllstoreName = (AllstoreName[dash + 1 :]).strip()
                    ListStoreName = AllstoreName.split()
                    storeFirstName = ListStoreName[0].lower()
                # print(storeFirstName)

        if isinstance(sheet["C4"].value, str):
            if sheet["C4"].value.startswith("BA"):
                BAVal1 = (sheet["C4"].value[9:]).strip().lower()
                BAVal2 = BAVal1.split("/")
                if len(BAVal2) > 1:
                    instances = len(BAVal2)

        excelSheet = classes.Excel(name, storeFirstName, BAVal1, instances)
        AllSheets.append(excelSheet)
    return AllSheets


def getMessages1(chatName, uptoDate, check):
    bot = Whatsapp(silent=True, headless=check)
    bot.login()
    messages = bot.getMessages(chatName, scroll=2)
    AllWhatsappInstances = []
    for msg in messages:
        instanceWA = 1
        AllMsg = []
        if str(msg[2]).startswith("Jam-E-Shirin"):
            message = msg[2]
            index = message.find("Date")
            message = message[index:].lstrip()
            bits = message.split("\n")
            # print(bits[0])
            colon = ":"
            index = bits[0].find(colon)
            # print("index: ", index)
            dateTemp = bits[0][index + 1 :].lstrip()
            date = dateTemp[:2] + "/" + dateTemp[3:5] + "/" + dateTemp[6:]
            dateFormat = dt.strptime(date, "%d/%m/%Y")
            checkDate = dt.strptime(uptoDate, "%d/%m/%Y")
            if dateFormat >= checkDate:
                # print(date)s
                index = bits[2].find(colon)
                BAName = bits[2][index + 1 :].lstrip()
                BANameLower = BAName.lower()
                ListBANamesSlash = BANameLower.split("/")
                ListBANamesSpace = BANameLower.split()
                BAFirstName = ListBANamesSpace[0].strip()
                if len(ListBANamesSlash) > 1:
                    instanceWA = len(ListBANamesSlash)
                index = bits[1].find(colon)
                store = bits[1][index + 1 :].lstrip()
                store = store.lower()
                # print(store)
                totalInterceptions = int(re.sub(r"[^0-9]", "", bits[4]))
                # print(totalInterceptions)
                totalSales = int(re.sub(r"[^0-9]", "", bits[5]))
                # print(totalSales)
                shortage = False
                for i in range(len(bits)):
                    if "ortage" in bits[i]:
                        index = i
                if "yes" in bits[i].lower():
                    shortage = True
                JamShirinSmall = 0
                JamShirinMedium = 0
                JamShirinLarge = 0
                JamShirinSugarFree = 0
                sandaleen = 0
                bazooreen = 0
                ilacheen = 0

                for i in range(len(bits)):
                    if "rin 800ml" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                JamShirinSmall = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "rin 1500ml" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                JamShirinMedium = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "rin 3000ml" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                JamShirinLarge = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "free" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("free")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                JamShirinSugarFree = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "aleen" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                sandaleen = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "oreen" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                bazooreen = int(re.sub(r"[^0-9]", "", ans))

                for i in range(len(bits)):
                    if "cheen" in bits[i]:
                        temp = i
                mlIndex = bits[temp].find("ml")
                ans = bits[temp][mlIndex + 1 :].lstrip()
                ilacheen = int(re.sub(r"[^0-9]", "", ans))

                AllMsg = [
                    date,
                    store,
                    totalInterceptions,
                    totalSales,
                    shortage,
                    BAFirstName,
                    JamShirinSmall,
                    JamShirinMedium,
                    JamShirinLarge,
                    JamShirinSugarFree,
                    sandaleen,
                    bazooreen,
                    ilacheen,
                    instanceWA,
                ]
                AllWhatsappInstances.append(AllMsg)
    return AllWhatsappInstances


def storeDate(messages, response, workBook, sheetName):
    try:
        book = load_workbook(workBook)
    except:
        print("File not found!")
        return 0
    if response == 1:
        try:
            # print("1")
            sheet = book[sheetName]
            # print("2")
            for msg in messages:
                tempmsg = msg[: len(msg) - 1]
                sheet.append(tempmsg)
                # print("3")
            book.save(workBook)
            # print("4")
            book.close()
            # print("5")
        except:
            print("sheet not found!")
            return 0
    else:
        doe = retrieveExcelInfo()
        for msg in messages:
            for itr in doe:
                if itr.store in msg[1]:
                    # print("inside store")
                    if msg[5].lower() in itr.BADetails.name.lower():
                        # print("inside name")
                        if int(itr.BADetails.instances) == int(msg[13]):
                            # print("inside instance")
                            dateDiff = (
                                dt.strptime(msg[0], "%d/%m/%Y")
                                - dt.strptime("14/03/2023", "%d/%m/%Y")
                            ).days
                            # print("Date diff: ", dateDiff)
                            column = column_names[5]
                            temp = 0

                            interceptions = msg[2]
                            cell = column + "7"
                            sheet = book[itr.sheetName]
                            sheet[cell] = interceptions

                            JamShirinSmall = msg[6]
                            cell = column + "8"
                            # print("sheetName: ", itr.sheetName)
                            sheet[cell] = JamShirinSmall
                            # print("tempVal: ", tempVal)

                            JamShirinMedium = msg[7]
                            cell = column + "9"
                            sheet[cell] = JamShirinMedium

                            JamShirinLarge = msg[8]
                            cell = column + "10"
                            sheet[cell] = JamShirinLarge

                            JamShirinSugarFree = msg[9]
                            cell = column + "11"
                            sheet[cell] = JamShirinSugarFree

                            sandaleen = msg[10]
                            cell = column + "12"
                            sheet[cell] = sandaleen

                            bazooreen = msg[11]
                            cell = column + "13"
                            sheet[cell] = bazooreen

                            ilacheen = msg[12]
                            cell = column + "14"
                            sheet[cell] = ilacheen
                            book.save(workBook)
    return 1


def is_valid_datetime(string):
    try:
        dt.strptime(string, "%d/%m/%Y")
        return True
    except ValueError:
        return False
